from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.rab import get_activity_state


def _currency(number: float) -> str:
    return f"Rp{number:,.0f}".replace(",", ".")


def build_excel_export(connection, activity_id: str) -> bytes:
    state = get_activity_state(connection, activity_id)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Ringkasan"

    header_fill = PatternFill("solid", fgColor="0E4F5C")
    subheader_fill = PatternFill("solid", fgColor="E3F1EF")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    bold_font = Font(bold=True)

    summary_sheet["A1"] = "RAB Workflow Assistant"
    summary_sheet["A2"] = state["activity"]["name"]
    summary_sheet["A1"].font = title_font
    summary_sheet["A1"].fill = header_fill
    summary_sheet["A2"].font = Font(bold=True, size=16)
    summary_sheet.merge_cells("A1:D1")
    summary_sheet.merge_cells("A2:D2")

    meta_rows = [
        ("Tahun Anggaran", state["activity"]["fiscal_year"]),
        ("Pagu Kegiatan", state["summary"]["budget_ceiling"]),
        ("Total RAB", state["summary"]["grand_total"]),
        ("Sisa Pagu", state["summary"]["remaining_budget"]),
    ]
    start_row = 4
    for index, (label, value) in enumerate(meta_rows, start=start_row):
        summary_sheet[f"A{index}"] = label
        summary_sheet[f"B{index}"] = value
        summary_sheet[f"A{index}"].font = bold_font
    for cell in ["B5", "B6", "B7"]:
        summary_sheet[cell].number_format = '#,##0'

    summary_sheet["A10"] = "Total per Sub Komponen"
    summary_sheet["A10"].fill = subheader_fill
    summary_sheet["A10"].font = bold_font
    row_cursor = 11
    for item in state["summary"]["totals_by_sub_component"]:
        summary_sheet[f"A{row_cursor}"] = f'{item["code"]}. {item["name"]}'
        summary_sheet[f"B{row_cursor}"] = item["total"]
        summary_sheet[f"B{row_cursor}"].number_format = '#,##0'
        row_cursor += 1

    row_cursor += 1
    summary_sheet[f"A{row_cursor}"] = "Total per Akun"
    summary_sheet[f"A{row_cursor}"].fill = subheader_fill
    summary_sheet[f"A{row_cursor}"].font = bold_font
    row_cursor += 1
    for item in state["summary"]["totals_by_account"]:
        summary_sheet[f"A{row_cursor}"] = f'{item["account_code"]} - {item["account_name"]}'
        summary_sheet[f"B{row_cursor}"] = item["total"]
        summary_sheet[f"B{row_cursor}"].number_format = '#,##0'
        row_cursor += 1

    detail_sheet = workbook.create_sheet("RAB Detail")
    headers = [
        "Kode Sub",
        "Sub Komponen",
        "Akun",
        "Nama Akun",
        "Detail Belanja",
        "Spesifikasi",
        "Volume",
        "Satuan",
        "Harga Satuan",
        "Subtotal",
        "Keterangan SBM",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = detail_sheet.cell(row=1, column=col_index)
        cell.value = header
        cell.fill = header_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    detail_row = 2
    for sub_component in state["sub_components"]:
        for account in sub_component["accounts"]:
            if not account["is_selected"]:
                continue
            for line in account["lines"]:
                detail_sheet.cell(row=detail_row, column=1, value=sub_component["code"])
                detail_sheet.cell(row=detail_row, column=2, value=sub_component["name"])
                detail_sheet.cell(row=detail_row, column=3, value=account["account_code"])
                detail_sheet.cell(row=detail_row, column=4, value=account["account_name"])
                detail_sheet.cell(row=detail_row, column=5, value=line["item_name"])
                detail_sheet.cell(row=detail_row, column=6, value=line["specification"])
                detail_sheet.cell(row=detail_row, column=7, value=line["volume"])
                detail_sheet.cell(row=detail_row, column=8, value=line["unit"])
                detail_sheet.cell(row=detail_row, column=9, value=line["unit_price"])
                detail_sheet.cell(row=detail_row, column=10, value=line["amount"])
                detail_sheet.cell(row=detail_row, column=11, value=line["suggestion_note"])
                detail_sheet.cell(row=detail_row, column=9).number_format = '#,##0'
                detail_sheet.cell(row=detail_row, column=10).number_format = '#,##0'
                detail_row += 1

    widths = {
        "A": 12,
        "B": 36,
        "C": 12,
        "D": 34,
        "E": 30,
        "F": 28,
        "G": 10,
        "H": 16,
        "I": 16,
        "J": 18,
        "K": 48,
    }
    for column_letter, width in widths.items():
        summary_sheet.column_dimensions[column_letter].width = width if column_letter in {"A", "B"} else summary_sheet.column_dimensions[column_letter].width
        detail_sheet.column_dimensions[column_letter].width = width

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_pdf_export(connection, activity_id: str) -> bytes:
    state = get_activity_state(connection, activity_id)
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=16, textColor=colors.HexColor("#0E4F5C"))
    subtitle_style = ParagraphStyle("subtitle", parent=styles["BodyText"], fontSize=9, leading=12)

    story = [
        Paragraph("RAB Workflow Assistant", title_style),
        Paragraph(state["activity"]["name"], styles["Heading2"]),
        Paragraph(
            f'Tahun Anggaran {state["activity"]["fiscal_year"]} | Pagu {_currency(state["summary"]["budget_ceiling"])} | Total {_currency(state["summary"]["grand_total"])} | Sisa {_currency(state["summary"]["remaining_budget"])}',
            subtitle_style,
        ),
        Spacer(1, 0.4 * cm),
    ]

    summary_table = Table(
        [["Sub Komponen", "Total"]] + [[f'{item["code"]}. {item["name"]}', _currency(item["total"])] for item in state["summary"]["totals_by_sub_component"]],
        colWidths=[18 * cm, 7 * cm],
        repeatRows=1,
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E4F5C")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7D0CD")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4FAF9")]),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 0.4 * cm)])

    detail_rows = [["Kode", "Sub Komponen", "Akun", "Detail", "Vol", "Sat", "Harga", "Subtotal"]]
    for sub_component in state["sub_components"]:
        for account in sub_component["accounts"]:
            if not account["is_selected"]:
                continue
            for line in account["lines"]:
                detail_rows.append(
                    [
                        sub_component["code"],
                        sub_component["name"],
                        account["account_code"],
                        line["item_name"],
                        str(line["volume"]),
                        line["unit"],
                        _currency(line["unit_price"]),
                        _currency(line["amount"]),
                    ]
                )

    detail_table = Table(detail_rows, colWidths=[1.4 * cm, 6 * cm, 2 * cm, 7.4 * cm, 1.7 * cm, 2 * cm, 3.2 * cm, 3.2 * cm], repeatRows=1)
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCEEEB")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C8D7D4")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBFB")]),
            ]
        )
    )
    story.append(detail_table)

    document.build(story)
    return buffer.getvalue()
