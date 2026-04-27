from __future__ import annotations

import json
import shutil
import sqlite3
import unicodedata
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.config import DATA_DIR


def normalize_key(value: str | None) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value)).upper().strip()
    return "".join(character for character in text if character.isalnum())


def _numeric(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _row_has_values(row: tuple[object, ...], min_index: int = 0) -> bool:
    return any(cell not in (None, "") for cell in row[min_index:])


def _insert_reference(connection: sqlite3.Connection, record: dict[str, object]) -> None:
    connection.execute(
        """
        INSERT INTO cost_references (
            id, source_sheet, category, label, region, region_key, origin, origin_key,
            destination, destination_key, unit, rate_primary, rate_secondary, rate_tertiary,
            rate_quaternary, meta_json
        )
        VALUES (
            :id, :source_sheet, :category, :label, :region, :region_key, :origin, :origin_key,
            :destination, :destination_key, :unit, :rate_primary, :rate_secondary, :rate_tertiary,
            :rate_quaternary, :meta_json
        )
        """,
        record,
    )


def _resolve_workbook_copy(workbook_path: Path) -> Path:
    cache_path = DATA_DIR / "sbm-cache.xlsx"
    if workbook_path.exists():
        try:
            shutil.copy2(workbook_path, cache_path)
            return cache_path
        except OSError:
            if cache_path.exists():
                return cache_path
            return workbook_path
    return cache_path if cache_path.exists() else workbook_path


def import_sbm_workbook(connection: sqlite3.Connection, workbook_path: Path) -> int:
    workbook = load_workbook(_resolve_workbook_copy(workbook_path), data_only=True)
    connection.execute("DELETE FROM cost_references")

    records: list[dict[str, object]] = []

    def add_record(
        *,
        source_sheet: str,
        category: str,
        label: str | None = None,
        region: str | None = None,
        origin: str | None = None,
        destination: str | None = None,
        unit: str | None = None,
        rate_primary: float | None = None,
        rate_secondary: float | None = None,
        rate_tertiary: float | None = None,
        rate_quaternary: float | None = None,
        meta: dict[str, object] | None = None,
    ) -> None:
        records.append(
            {
                "id": uuid4().hex,
                "source_sheet": source_sheet,
                "category": category,
                "label": label,
                "region": region,
                "region_key": normalize_key(region),
                "origin": origin,
                "origin_key": normalize_key(origin),
                "destination": destination,
                "destination_key": normalize_key(destination),
                "unit": unit,
                "rate_primary": rate_primary,
                "rate_secondary": rate_secondary,
                "rate_tertiary": rate_tertiary,
                "rate_quaternary": rate_quaternary,
                "meta_json": json.dumps(meta or {}, ensure_ascii=True),
            }
        )

    honor = workbook["HONOR"]
    for row in honor.iter_rows(min_row=3, values_only=True):
        if not _row_has_values(row, 1) or row[1] is None:
            continue
        add_record(
            source_sheet="HONOR",
            category="honorarium",
            label=str(row[1]).strip(),
            unit=row[2],
            rate_primary=_numeric(row[3]),
        )

    daily_allowance = workbook["UH DN"]
    for row in daily_allowance.iter_rows(min_row=3, values_only=True):
        province = row[1]
        if province in (None, ""):
            continue
        add_record(
            source_sheet="UH DN",
            category="daily_allowance_domestic",
            region=str(province).strip(),
            unit=row[2],
            rate_primary=_numeric(row[3]),
            rate_secondary=_numeric(row[4]),
            rate_tertiary=_numeric(row[5]),
        )

    hotel = workbook["HOTEL"]
    for row in hotel.iter_rows(min_row=4, values_only=True):
        province = row[1]
        if province in (None, ""):
            continue
        add_record(
            source_sheet="HOTEL",
            category="hotel_domestic",
            region=str(province).strip(),
            unit=row[2],
            rate_primary=_numeric(row[3]),
            rate_secondary=_numeric(row[4]),
            rate_tertiary=_numeric(row[5]),
            rate_quaternary=_numeric(row[6]),
        )

    meeting_package = workbook["PM"]
    for row in meeting_package.iter_rows(min_row=3, values_only=True):
        province = row[1]
        if province in (None, ""):
            continue
        add_record(
            source_sheet="PM",
            category="meeting_package",
            region=str(province).strip(),
            unit=row[2],
            rate_primary=_numeric(row[3]),
            rate_secondary=_numeric(row[4]),
            rate_tertiary=_numeric(row[5]),
        )

    sewa = workbook["SEWA"]
    for row in sewa.iter_rows(min_row=3, values_only=True):
        province = row[1]
        if province in (None, ""):
            continue
        add_record(
            source_sheet="SEWA",
            category="vehicle_rent",
            region=str(province).strip(),
            unit=row[2],
            rate_primary=_numeric(row[3]),
            rate_secondary=_numeric(row[4]),
            rate_tertiary=_numeric(row[5]),
        )

    konsum = workbook["KONSUM"]
    for row in konsum.iter_rows(min_row=5, values_only=True):
        province = row[1]
        if province in (None, ""):
            continue
        add_record(
            source_sheet="KONSUM",
            category="consumption_regular",
            region=str(province).strip(),
            unit=row[2],
            rate_primary=_numeric(row[3]),
            rate_secondary=_numeric(row[4]),
        )

    ground_transport = workbook["TRANS DARAT"]
    current_province: str | None = None
    for row in ground_transport.iter_rows(min_row=1, values_only=True):
        title = str(row[0]).strip() if row[0] not in (None, "") else ""
        column_b = str(row[1]).strip() if row[1] not in (None, "") else ""

        if "SATUAN BIAYA TRANSPORTASI DARI DKI JAKARTA" in title.upper():
            current_province = "DKI JAKARTA"
            continue

        if "SATUAN BIAYA TRANSPOR KEGIATAN DALAM KABUPATEN/KOTA PERGI PULANG (PP)" in title.upper():
            add_record(
                source_sheet="TRANS DARAT",
                category="local_transport_activity_pp",
                label="Transport kegiatan dalam kabupaten/kota PP",
                unit=row[3],
                rate_primary=_numeric(row[4]),
            )
            current_province = None
            continue

        if row[0] in (None, "") and row[1] not in (None, "") and row[2] in (None, ""):
            current_province = column_b
            continue
        if row[0] in (None, "") or row[2] in (None, ""):
            continue
        add_record(
            source_sheet="TRANS DARAT",
            category="land_transport_domestic",
            region=current_province,
            origin=str(row[1]).strip(),
            destination=str(row[2]).strip(),
            unit=row[3],
            rate_primary=_numeric(row[4]),
        )

    airport_taxi = workbook["TAKSI BANDARA"]
    for row in airport_taxi.iter_rows(min_row=3, values_only=True):
        province = row[1]
        if province in (None, ""):
            continue
        add_record(
            source_sheet="TAKSI BANDARA",
            category="airport_taxi_domestic",
            region=str(province).strip(),
            unit=row[2],
            rate_primary=_numeric(row[3]),
        )

    flights = workbook["PESAWAT DN"]
    for row in flights.iter_rows(min_row=4, values_only=True):
        if row[0] in (None, "") or row[2] in (None, ""):
            continue
        add_record(
            source_sheet="PESAWAT DN",
            category="flight_domestic_pp",
            origin=str(row[1]).strip(),
            destination=str(row[2]).strip(),
            unit="PP",
            rate_primary=_numeric(row[3]),
            rate_secondary=_numeric(row[4]),
        )

    for record in records:
        _insert_reference(connection, record)

    return len(records)


def seed_fallback_costs(connection: sqlite3.Connection) -> int:
    connection.execute("DELETE FROM cost_references")
    fallback_records = [
        {
            "id": uuid4().hex,
            "source_sheet": "FALLBACK",
            "category": "consumption_regular",
            "label": "Fallback konsumsi rapat biasa",
            "region": "DKI JAKARTA",
            "region_key": normalize_key("DKI JAKARTA"),
            "origin": None,
            "origin_key": "",
            "destination": None,
            "destination_key": "",
            "unit": "Orang/Kali",
            "rate_primary": 60000,
            "rate_secondary": 25000,
            "rate_tertiary": None,
            "rate_quaternary": None,
            "meta_json": json.dumps({"fallback": True}),
        },
        {
            "id": uuid4().hex,
            "source_sheet": "FALLBACK",
            "category": "honorarium",
            "label": "Honorarium Narasumber",
            "region": None,
            "region_key": "",
            "origin": None,
            "origin_key": "",
            "destination": None,
            "destination_key": "",
            "unit": "OJ",
            "rate_primary": 1000000,
            "rate_secondary": None,
            "rate_tertiary": None,
            "rate_quaternary": None,
            "meta_json": json.dumps({"fallback": True}),
        },
        {
            "id": uuid4().hex,
            "source_sheet": "FALLBACK",
            "category": "daily_allowance_domestic",
            "label": None,
            "region": "DKI JAKARTA",
            "region_key": normalize_key("DKI JAKARTA"),
            "origin": None,
            "origin_key": "",
            "destination": None,
            "destination_key": "",
            "unit": "OH",
            "rate_primary": 530000,
            "rate_secondary": 230000,
            "rate_tertiary": 150000,
            "rate_quaternary": None,
            "meta_json": json.dumps({"fallback": True}),
        },
        {
            "id": uuid4().hex,
            "source_sheet": "FALLBACK",
            "category": "hotel_domestic",
            "label": None,
            "region": "DKI JAKARTA",
            "region_key": normalize_key("DKI JAKARTA"),
            "origin": None,
            "origin_key": "",
            "destination": None,
            "destination_key": "",
            "unit": "OH",
            "rate_primary": 4800000,
            "rate_secondary": 2500000,
            "rate_tertiary": 1700000,
            "rate_quaternary": 950000,
            "meta_json": json.dumps({"fallback": True}),
        },
    ]
    connection.executemany(
        """
        INSERT INTO cost_references (
            id, source_sheet, category, label, region, region_key, origin, origin_key,
            destination, destination_key, unit, rate_primary, rate_secondary, rate_tertiary,
            rate_quaternary, meta_json
        )
        VALUES (
            :id, :source_sheet, :category, :label, :region, :region_key, :origin, :origin_key,
            :destination, :destination_key, :unit, :rate_primary, :rate_secondary, :rate_tertiary,
            :rate_quaternary, :meta_json
        )
        """,
        fallback_records,
    )
    return len(fallback_records)
